import os
import shutil
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows
import re
from typing import List, Tuple
from openpyxl.utils import get_column_letter
import argparse



def combine_fastas(folder_paths):
    """
    Combine FASTA files from multiple folders into a single directory.

    Parameters:
        folder_paths (list[str]): List of paths containing FASTA files.

    Returns:
        str: Path to the combined "final_extracted_fastas" directory.
    """
    base_dir = os.path.dirname(os.path.abspath(__file__))
    final_dir = os.path.join(base_dir, "final_extracted_fastas")
    os.makedirs(final_dir, exist_ok=True)

    total_copied = 0
    total_removed = 0

    for folder in folder_paths:
        if not os.path.isdir(folder):
            print(f"[WARN] Skipping non-directory: {folder}")
            continue

        for file in os.listdir(folder):
            if not file.lower().endswith((".fasta", ".fa", ".faa")):
                continue

            src_path = os.path.join(folder, file)
            if os.path.getsize(src_path) == 0:
                print(f"[REMOVE] 0-byte file: {src_path}")
                total_removed += 1
                continue

            new_name = f"{file}"
            dest_path = os.path.join(final_dir, new_name)

            shutil.copy2(src_path, dest_path)
            total_copied += 1

    print(f"\nFASTA combination complete.")
    print(f"Total files copied: {total_copied}")
    print(f"0-byte files skipped: {total_removed}")
    print(f"Output directory: {final_dir}")

    return final_dir



def merge_summary_excels(excel_paths: List[str], output_path: str = "merged_summary.xlsx") -> str:
    """
    Merge multiple 'BLAST Summary' Excel sheets into one formatted workbook.

    Fixes:
    - Proper category row (row 1)
    - Clean headers (no Unnamed/.1)
    - Recomputed colors + r,y,r+y
    - Summary block not colored
    - Skips blank/summary rows in color loop
    - Freezes panes at C3
    """

    # ---- Formatting ----
    RED_FILL = PatternFill("solid", fgColor="FF6347")
    YELLOW_FILL = PatternFill("solid", fgColor="FFD700")
    BLUE_FILL = PatternFill("solid", fgColor="87CEFA")
    HEADER_FONT = Font(bold=True)
    CENTER = Alignment(horizontal="center")
    EXCLUDED = {"Organism", "A-score", "r,y,r+y", "Phylum", "Source_File"}

    FEATURES = ["AMR", "Bio", "Hpat", "TH", "RAD", "SF"]
    PROTEIN_CATEGORY = {
        "AMR": ["ERG11", "CDR1", "MDR1", "UPC2", "TAC1", "MRR1"],
        "Bio": ["BCR1", "EFG1", "TEC1", "HWP1", "ALS3", "NDT80", "ERG251", "CZF1", "FLO8"],
        "Hpat": ["SAP5", "PLB1", "LAC1", "RIM101"],
        "TH": ["HSP90"],
        "RAD": ["RAD51"],
        "SF": ["brlA", "abaA", "wetA", "srr1"],
    }

    # ---- Helper functions ----
    def clean_columns(cols):
        out = []
        for c in cols:
            if isinstance(c, str) and c.startswith("Unnamed:"):
                continue
            if isinstance(c, str):
                c = re.sub(r"\.\d+$", "", c)
            out.append(c)
        return out

    def read_one(path):
        df = pd.read_excel(path, header=1, engine="openpyxl")
        df = df.iloc[:-11]
        df.columns = clean_columns(df.columns)
        if "Organism" in df.columns:
            df = df[~df["Organism"].isna()]
        df["Source_File"] = os.path.basename(path)
        return df

    def protein_columns(df):
        cols = []
        for c in df.columns:
            if c in EXCLUDED:
                continue
            if pd.api.types.is_numeric_dtype(df[c]) or pd.to_numeric(df[c], errors="coerce").notna().any():
                cols.append(c)
        return cols

    def row_counts(row, pcols):
        red = yellow = 0
        for c in pcols:
            try:
                v = float(row[c])
            except Exception:
                continue
            if v > 75:
                red += 1
            elif v > 35:
                yellow += 1
        return red, yellow

    def category_for_protein(prot_name: str) -> str:
        pl = prot_name.lower()
        for cat in FEATURES:
            for abbr in PROTEIN_CATEGORY.get(cat, []):
                if abbr.lower() in pl:
                    return cat
        return "Uncategorized"

    # ---- Read all valid inputs ----
    frames = []
    for path in excel_paths:
        if not os.path.exists(path):
            print(f"[WARN] Missing file: {path}")
            continue
        try:
            frames.append(read_one(path))
        except Exception as e:
            print(f"[WARN] Could not read 'BLAST Summary' in {path}: {e}")

    if not frames:
        print("No valid Excel files found.")
        return None

    merged = pd.concat(frames, ignore_index=True, sort=False)
    pcols = protein_columns(merged)

    # Column order
    ordered = []
    if "Organism" in merged.columns: ordered.append("Organism")
    if "A-score" in merged.columns: ordered.append("A-score")
    ordered += pcols
    if "r,y,r+y" in merged.columns: ordered.append("r,y,r+y")
    if "Phylum" in merged.columns: ordered.append("Phylum")
    if "Source_File" in merged.columns: ordered.append("Source_File")
    merged = merged[ordered]

    # Recompute r,y,r+y
    ry = merged.apply(lambda r: row_counts(r, pcols), axis=1, result_type="expand")
    merged["r,y,r+y"] = (
        ry[0].astype(int).astype(str) + "," +
        ry[1].astype(int).astype(str) + "," +
        (ry[0] + ry[1]).astype(int).astype(str)
    )

    # ---- Write workbook ----
    wb = Workbook()
    ws = wb.active
    ws.title = "BLAST Summary"

    # Category row (row 1)
    ws.cell(row=1, column=1, value="")
    ws.cell(row=1, column=2, value="")
    col_idx = 3
    for prot in pcols:
        cat = category_for_protein(prot)
        cell = ws.cell(row=1, column=col_idx, value=cat)
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        col_idx += 1
    tail_cols = [c for c in merged.columns if c not in (["Organism", "A-score"] + pcols)]
    for _ in tail_cols:
        ws.cell(row=1, column=col_idx, value="")
        col_idx += 1

    # Header row (row 2)
    for j, c in enumerate(merged.columns, start=1):
        cell = ws.cell(row=2, column=j, value=c)
        cell.font = HEADER_FONT
        cell.alignment = CENTER

    # Data rows (start row 3)
    start_row = 3
    for i, (_, row) in enumerate(merged.iterrows(), start=start_row):
        org = str(row.get("Organism", "")).strip()
        if not org or "summary statistics" in org.lower():
            # skip blank or accidental summary rows
            continue

        for j, col in enumerate(merged.columns, start=1):
            val = row[col]
            cell = ws.cell(row=i, column=j, value=val)
            if col in pcols:
                try:
                    v = float(val)
                except Exception:
                    continue
                if v > 75:
                    cell.fill = RED_FILL
                elif v > 35:
                    cell.fill = YELLOW_FILL
                else:
                    cell.fill = BLUE_FILL
                cell.alignment = CENTER

        # color only valid organism cells
        r_cnt, y_cnt = row_counts(row, pcols)
        org_cell = ws.cell(row=i, column=1)
        if r_cnt > 0:
            org_cell.fill = RED_FILL
        elif y_cnt > 0:
            org_cell.fill = YELLOW_FILL
        else:
            org_cell.fill = BLUE_FILL

    # Adjust widths
    for col_cells in ws.columns:
        letter = col_cells[0].column_letter
        max_len = max((len(str(c.value)) if c.value else 0) for c in col_cells[:5000])
        ws.column_dimensions[letter].width = min(max_len + 2, 40)

    ws.freeze_panes = "C3"

    # ---- Summary block (uncolored) ----
    n = len(merged)
    red_rows = yellow_only = 0
    for _, r in merged.iterrows():
        rc, yc = row_counts(r, pcols)
        if rc > 0:
            red_rows += 1
        elif yc > 0:
            yellow_only += 1
    blue_rows = n - red_rows - yellow_only

    pct_red = round((red_rows / n) * 100, 1) if n else 0
    pct_yellow = round((yellow_only / n) * 100, 1) if n else 0
    pct_blue = round((blue_rows / n) * 100, 1) if n else 0

    s = start_row + n + 1
    ws.cell(row=s, column=1, value="Summary Statistics").font = Font(bold=True, underline="single")
    ws.cell(row=s+1, column=1, value="Total organisms"); ws.cell(row=s+1, column=2, value=n)
    ws.cell(row=s+2, column=1, value="Rows with ≥1 red cell"); ws.cell(row=s+2, column=2, value=red_rows)
    ws.cell(row=s+3, column=1, value="Rows with ≥1 yellow (no red)"); ws.cell(row=s+3, column=2, value=yellow_only)
    ws.cell(row=s+4, column=1, value="Completely blue rows"); ws.cell(row=s+4, column=2, value=blue_rows)
    ws.cell(row=s+6, column=1, value="% with red"); ws.cell(row=s+6, column=2, value=pct_red)
    ws.cell(row=s+7, column=1, value="% with yellow"); ws.cell(row=s+7, column=2, value=pct_yellow)
    ws.cell(row=s+8, column=1, value="% completely blue"); ws.cell(row=s+8, column=2, value=pct_blue)

    # Clear any leftover fills below data
    for r in ws.iter_rows(min_row=start_row + len(merged), max_row=ws.max_row):
        for c in r:
            c.fill = PatternFill()

    wb.save(output_path)
    print(f"Excel merge complete → {output_path}")
    return output_path




if __name__ == "__main__":

    parser = argparse.ArgumentParser(description="Merge multiple 'BLAST Summary' Excel sheets or combine FASTA files.")
    parser.add_argument("--excels", nargs="+", help="List of Excel files to merge.", default=None)
    parser.add_argument("--output", default="merged_summary.xlsx", help="Output Excel file path.")
    parser.add_argument("--fastas", nargs="+", help="List of folders containing FASTA files to combine.", default=None)
    args = parser.parse_args()

    if args.excels:
        merge_summary_excels(args.excels, args.output)
    elif args.fastas:
        combine_fastas(args.fastas)
    else:
        print("Please provide either --excels or --fastas arguments.")



