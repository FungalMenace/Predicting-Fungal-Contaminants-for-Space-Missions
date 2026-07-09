"""
make_summary_excel.py
---------------------
Generates a color-coded Excel summary of BLAST identity results.

Input:
    A folder containing one CSV per organism (e.g., "Candida_albicans.csv").
    Each CSV must contain two columns: [Protein, Identity].

Output:
    Excel workbook summarizing all organisms × proteins,
    color-coded by thresholds and including per-organism A-scores
    and Phylum information.

Usage:
    python make_summary_excel.py --i ./blast_results --o fungal_summary.xlsx
"""

import os
import glob
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
import argparse
from pathlib import Path
from src.fungi_pipeline.excel.phyla import get_phylum 


FEATURES = [
    "AMR",
    "Biofilm",
    "H-pat",
    "Thermophile",
    "Rad-res",
    "Spore",
]

PROTEIN_CATEGORY = {
    "AMR": [
        "ERG11",
        "CDR1",
        "MDR1",
        "UPC2",
        "TAC1",
        "MRR1",
    ],

    "Biofilm": [
        "BCR1",
        "EFG1",
        "TEC1",
        "HWP1",
        "ALS3",
        "NDT80",
        "ERG251",
        "CZF1",
        "FLO8",
    ],

    "H-pat": [
        "SAP5",
        "PLB1",
        "LAC1",
        "RIM101",
    ],

    "Thermophile": [
        "HSP90",
    ],

    "Rad-res": [
        "RAD51",
    ],

    "Spore": [
        "brlA",
        "abaA",
        "wetA",
        "srr1",
    ],
}

HEADER_INFO = [
    ("AMR", "ERG11", "Candida", "Albicans"),
    ("AMR", "CDR1", "Candida", "Albicans"),
    ("AMR", "MDR1", "Candida", "Albicans"),
    ("AMR", "UPC2", "Candida", "Albicans"),
    ("AMR", "TAC1", "Candida", "Albicans"),
    ("AMR", "MRR1", "Candida", "Albicans"),

    ("Biofilm", "BCR1", "Candida", "Albicans"),
    ("Biofilm", "EFG1", "Candida", "Albicans"),
    ("Biofilm", "TEC1", "Candida", "Albicans"),
    ("Biofilm", "HWP1", "Candida", "Albicans"),
    ("Biofilm", "ALS3", "Candida", "Albicans"),
    ("Biofilm", "NDT80", "Candida", "Albicans"),
    ("Biofilm", "ERG251", "Candida", "Albicans"),
    ("Biofilm", "CZF1", "Candida", "Albicans"),
    ("Biofilm", "FLO8", "Candida", "Albicans"),

    ("H-pat", "SAP5", "Candida", "Albicans"),
    ("H-pat", "PLB1", "Cryptococcus", "Neoformans"),
    ("H-pat", "LAC1", "Cryptococcus", "Neoformans"),
    ("H-pat", "RIM101", "Candida", "Albicans"),

    ("Thermophile", "HSP90", "Aspergillus", "Fumigatus"),

    ("Rad-res", "RAD51", "Saccharomyces", "Cerevisiae"),

    ("Spore", "brlA", "Emericella", "Nidulans"),
    ("Spore", "abaA", "Emericella", "Nidulans"),
    ("Spore", "wetA", "Emericella", "Nidulans"),
    ("Spore", "srr1", "Coprinopsis", "Cinerea"),
]


RED_FILL = PatternFill("solid", fgColor="FF6347")     # >75
YELLOW_FILL = PatternFill("solid", fgColor="FFD700")  # >35
BLUE_FILL = PatternFill("solid", fgColor="87CEFA")    # ≤35



def read_blast_results(folder):
    """Return (data, organisms, proteins)."""
    csvs = sorted(glob.glob(os.path.join(folder, "*.csv")))
    if not csvs:
        raise FileNotFoundError(f"No CSV files found in {folder}")

    data = {}
    all_proteins = set()
    organisms = []

    for path in csvs:
        organism = os.path.basename(path).replace(".csv", "")
        organism = organism.split("_blast")[0].replace("_", " ")
        if organism == "all results":
            continue
        organisms.append(organism)

        df = pd.read_csv(path)
        protein_col, pident_col = "protein", "pident"
        for _, row in df.iterrows():
            prot = str(row[protein_col]).strip()
            try:
                val = float(row[pident_col])
            except Exception:
                continue
            data[(organism, prot)] = val
            all_proteins.add(prot)

    return data, sorted(set(organisms)), sorted(all_proteins)



def load_source_mapping(source_file):
    if source_file is None or not os.path.exists(source_file):
        return {}

    mapping = {}

    with open(source_file) as f:
        for line in f:
            line = line.strip()
            if not line:
                continue

            if "\t" in line:
                org, src = line.split("\t", 1)
            else:
                org, src = line.split(",", 1)

            mapping[org.strip()] = src.strip()

    return mapping

def load_phylum_cache(cache_path="phylum_cache.json"):
    if os.path.exists(cache_path):
        try:
            with open(cache_path, "r") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def save_phylum_cache(cache, cache_path="phylum_cache.json"):
    try:
        with open(cache_path, "w") as f:
            json.dump(cache, f, indent=2)
    except Exception as e:
        print(f"⚠️ Could not save phylum cache: {e}")



def generate_excel(
    data,
    organisms,
    all_proteins,
    output_path,
    source_file=None,
    cache_path="phylum_cache.json",
):
    source_map = load_source_mapping(source_file)
    # Map expected protein names to the actual BLAST protein names
    protein_lookup = {}

    for _, blast_protein in data.keys():
        blast_upper = blast_protein.upper()

        for _, expected, _, _ in HEADER_INFO:
            if blast_upper.startswith(expected.upper() + "-"):
                protein_lookup[expected] = blast_protein
    wb = Workbook()
    ws = wb.active
    ws.title = "BLAST Summary"

    ws.cell(row=2, column=1, value="Organism").font = Font(bold=True)
    ws.cell(row=2, column=2, value="A-score").font = Font(bold=True)

    for col, (cat, prot, genus, species) in enumerate(HEADER_INFO, start=3):
        ws.cell(row=1, column=col, value=cat)
        ws.cell(row=2, column=col, value=prot)
        ws.cell(row=3, column=col, value=genus)
        ws.cell(row=4, column=col, value=species)

        for r in range(1, 5):
            ws.cell(row=r, column=col).font = Font(bold=True)
            ws.cell(row=r, column=col).alignment = Alignment(horizontal="center")

    summary_col = len(HEADER_INFO) + 3
    source_col = summary_col + 1
    phylum_col = summary_col + 2

    ws.cell(row=2, column=summary_col, value="r,y,r+y").font = Font(bold=True)
    ws.cell(row=2, column=source_col, value="Source").font = Font(bold=True)
    ws.cell(row=2, column=phylum_col, value="Phyla").font = Font(bold=True)

    # Load cache
    phylum_cache = load_phylum_cache(cache_path)

    red_rows = 0
    yellow_rows = 0
    blue_rows = 0

    for i, org in enumerate(organisms, start=5):
        ws.cell(row=i, column=1, value=org)

        cat_red = {cat: 0 for cat in FEATURES}
        cat_yellow = {cat: 0 for cat in FEATURES}
        row_red = 0
        row_yellow = 0

        for j, (_, prot, _, _) in enumerate(HEADER_INFO, start=3):
            real_protein = protein_lookup.get(prot)

            if real_protein is None:
                val = 0.0
            else:
                val = data.get((org, real_protein), 0.0)
            c = ws.cell(row=i, column=j, value=round(val, 1))
            c.alignment = Alignment(horizontal="center")

            cat = next(c for c, p, _, _ in HEADER_INFO if p == prot)
            if cat in FEATURES:
                if val > 75:
                    cat_red[cat] += 1
                elif val > 35:
                    cat_yellow[cat] += 1

            if val > 75:
                c.fill = RED_FILL
                row_red += 1
            elif val > 35:
                c.fill = YELLOW_FILL
                row_yellow += 1
            else:
                c.fill = BLUE_FILL

        # A-score
        a_score = sum(2 if cat_red[c] > 0 else 1 if cat_yellow[c] > 0 else 0 for c in FEATURES)
        a_cell = ws.cell(row=i, column=2, value=a_score)
        a_cell.alignment = Alignment(horizontal="center")

        # Row color and classification
        hdr = ws.cell(row=i, column=1)
        if row_red > 0:
            hdr.fill = RED_FILL
            red_rows += 1
        elif row_yellow > 0:
            hdr.fill = YELLOW_FILL
            yellow_rows += 1
        else:
            hdr.fill = BLUE_FILL
            blue_rows += 1

        ws.cell(row=i, column=summary_col, value=f"{row_red},{row_yellow},{row_red + row_yellow}")

        ws.cell(row=i, column=source_col,
        value=source_map.get(org, ""))

        # --- Identify Phylum ---
        org_clean = org.strip()
        if org_clean in phylum_cache:
            phylum = phylum_cache[org_clean]
        else:
            phylum = get_phylum(org_clean)
            phylum_cache[org_clean] = phylum
            save_phylum_cache(phylum_cache, cache_path)  # Save incrementally
        ws.cell(row=i, column=phylum_col, value=phylum)


    ws.column_dimensions[get_column_letter(1)].width = 28
    ws.column_dimensions[get_column_letter(2)].width = 12
    for j in range(3, phylum_col + 1):
        ws.column_dimensions[get_column_letter(j)].width = 16
    ws.freeze_panes = "C5"


    total_rows = len(organisms)
    last_row = len(organisms) + 5
    ws.cell(row=last_row, column=1, value="Summary Statistics").font = Font(bold=True, underline="single")

    ws.cell(row=last_row + 1, column=1, value="Total organisms")
    ws.cell(row=last_row + 1, column=2, value=total_rows)

    ws.cell(row=last_row + 2, column=1, value="Rows with ≥1 red cell")
    ws.cell(row=last_row + 2, column=2, value=red_rows)

    ws.cell(row=last_row + 3, column=1, value="Rows with ≥1 yellow (no red)")
    ws.cell(row=last_row + 3, column=2, value=yellow_rows)

    ws.cell(row=last_row + 4, column=1, value="Completely blue rows")
    ws.cell(row=last_row + 4, column=2, value=blue_rows)

    # Percentages
    ws.cell(row=last_row + 6, column=1, value="% with red")
    ws.cell(row=last_row + 6, column=2, value=round((red_rows / total_rows) * 100, 1))

    ws.cell(row=last_row + 7, column=1, value="% with yellow")
    ws.cell(row=last_row + 7, column=2, value=round((yellow_rows / total_rows) * 100, 1))

    ws.cell(row=last_row + 8, column=1, value="% completely blue")
    ws.cell(row=last_row + 8, column=2, value=round((blue_rows / total_rows) * 100, 1))


    wb.save(output_path)
    print(f"Excel saved with Phylum and statistics: {output_path}")
    print(f"Cache updated at: {cache_path}")



def main():
    parser = argparse.ArgumentParser(description="Create Excel summary from BLAST result CSVs.")
    parser.add_argument("--i", default="src/results", help="Folder with organism CSV files.")
    parser.add_argument("--o", default="fungal_summary.xlsx", help="Output Excel file path.")
    args = parser.parse_args()

    data, organisms, prots = read_blast_results(args.i)
    generate_excel(
        data,
        organisms,
        prots,
        args.o,
        source_file=None, 
    )


if __name__ == "__main__":
    main()
