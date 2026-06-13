import os
import json
import glob
import pandas as pd
import requests
from io import StringIO
from Bio import SeqIO
from Bio.Blast.Applications import NcbiblastpCommandline
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
from src.fungi_pipeline.excel.phyla import get_phylum



SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_DIR = os.path.join(SCRIPT_DIR, "orthodb_results")
FASTAS_DIR = os.path.join(BASE_DIR, "fastas")
CSVS_DIR = os.path.join(BASE_DIR, "csvs")
TSVS_DIR = os.path.join(BASE_DIR, "tsvs")
GOLD_CSV = "/Users/vannshjani/Downloads/FungalMenaceToo/docs/gold_20250607_211424.csv"
EXCEL_PATH = os.path.join(BASE_DIR, "fungi_from_orthologs.xlsx")

for d in [BASE_DIR, FASTAS_DIR, CSVS_DIR, TSVS_DIR]:
    os.makedirs(d, exist_ok=True)


# -------------------- UNIPROT IDS --------------------

uniprot_ids = {
    "ERG11-Candida albicans": "P10613",
    "FSK2-Candida Glabrata": "Q6FMZ3",
    "CDR1-Candida albicans": "Q5ANA3",
    "MDR1-Candida albicans": "Q5ABU7",
    "PDR1-Candida Glabrata": "Q6FXU7",
    "UPC2-Candida albicans": "Q59QC7",
    "TAC1-Candida albicans": "A0A1D8PN96",
    "MRR1-Candida albicans": "Q5A4G2",
    "BCR1-Candida albicans": "Q59U10",
    "EFG1-Candida albicans": "Q59X67",
    "TEC1-Candida albicans": "Q5ANJ4",
    "HWP1-Candida albicans": "P46593",
    "ALS3-Candida albicans": "Q59L12",
    "NDT80-Candida albicans": "Q5ACU9",
    "FLO8-Candida albicans": "Q59QW5",
    "CZF1-Candida albicans": "Q5A0W9",
    "ERG251-Candida albicans": "A0A1D8PLB5",
    "SAP5-Candida albicans": "P43094",
    "PLB1-Cryptococcus neoformans": "Q9P8P2",
    "CAP59-Cryptococcus neoformans": "Q9URK1",
    "LAC1-Cryptococcus neoformans": "Q55P57",
    "RIM101-Candida albicans": "Q9UW14",
    "HSP90-Aspergillus fumigatus": "P40292",
    "TPS1-Cryptococcus neoformans": "Q6IVK9",
    "AFP4-Glaciozyma antarctica": "M1JFL3",
    "PKS1-Exophiala dermatitidis": "Q9Y7A7",
    "RAD51-Saccharomyces cerevisiae": "P25454",
    "ku70-Exophiala dermatitidis": "A0AAN6IZM1",
    "brlA-Emericella nidulans": "P10069",
    "abaA-Emericella nidulans": "P20945",
    "wetA-Emericella nidulans": "P22022",
    "srr1-Coprinopsis cinerea": "A8NH73",
}


# -------------------- CLASS --------------------

class FUNGALMENACE:
    def __init__(self, uniprot_id, file_name):
        self.uniprot_id = uniprot_id
        self.file_name = file_name.replace(" ", "-").replace("/", "-")
        self.og_id = None
        self.identity_thresh = 0
        self.coverage_thresh = 0
        self.query_length = 0
        self.ortho_file_name = f"{self.file_name}_orthologs.fasta"

    def get_orthodb_ogs(self):
        url = f"https://data.orthodb.org/v12/search?query={self.uniprot_id}"
        r = requests.get(url)
        r.raise_for_status()
        data = r.json()
        return data.get("bigdata", [])

    def fetch_ortholog_sequences(self):
        url = "https://data.orthodb.org/v12/fasta"
        params = {"id": self.og_id}
        r = requests.get(url, params=params, allow_redirects=True)
        r.raise_for_status()
        return list(SeqIO.parse(StringIO(r.text), "fasta"))

    def save_fasta(self, sequences, org=False):
        outpath = os.path.join(FASTAS_DIR, self.ortho_file_name if org else f"{self.file_name}.fasta")
        with open(outpath, "w") as f:
            SeqIO.write(sequences, f, "fasta")

    def fetch_query_sequence(self):
        url = f"https://rest.uniprot.org/uniprotkb/{self.uniprot_id}.fasta"
        outpath = os.path.join(FASTAS_DIR, f"{self.file_name}_query.fasta")
        with open(outpath, "w") as f:
            f.write(requests.get(url).text)
        return outpath

    def run_blastp(self, query_fasta, db_fasta, output_tsv):
        db_out = os.path.join(BASE_DIR, "orthodb_db")
        os.system(f"makeblastdb -in {db_fasta} -dbtype prot -out {db_out}")
        blastp_cline = NcbiblastpCommandline(
            query=query_fasta,
            db=db_out,
            evalue=1e-5,
            outfmt="6 qseqid sseqid pident length mismatch gapopen qstart qend sstart send evalue bitscore",
            out=output_tsv
        )
        blastp_cline()

    def is_probable_type_strain(self, species_name: str) -> bool:
        """
        Flag probable type strains based on common codes in the species name.
        """
        type_indicators = ["SC5314", "S288C", "Af293", "CBS", "NRRL",
                           "ATCC", "DSM", "IFO", "NBRC"]
        return any(tag in species_name for tag in type_indicators)

    def normalize_species(self, name: str) -> str:
        """
        Normalize species name by stripping strain identifiers (e.g., 'CBS 1234', 'NRRL Y-1111')
        """
        if pd.isna(name):
            return name
        name = name.strip()
        # Remove trailing strain indicators
        parts = name.split()
        if len(parts) >= 3 and parts[-2].isalpha() and parts[-1].replace(".", "").isdigit():
            return " ".join(parts[:2])
        # Handle 'sp.' variants like 'Tulasnella sp. 417'
        if "sp." in name:
            idx = name.index("sp.")
            return name[:idx+3].strip()
        return name

    def filter_blast_results(self, blast_tsv, query_length, fasta_file):
        cols = ["query","subject","identity","alignment_length","mismatches","gap_openings",
                "q_start","q_end","s_start","s_end","evalue","bit_score"]
        df = pd.read_csv(blast_tsv, sep="\t", header=None, names=cols)
        df["coverage"] = df["alignment_length"] / query_length
        df = df[(df["identity"] >= self.identity_thresh) & (df["coverage"] >= self.coverage_thresh)]

        seq_records = SeqIO.to_dict(SeqIO.parse(fasta_file, "fasta"))
        org_info = []
        for sid in df["subject"]:
            if sid in seq_records:
                try:
                    header_json = seq_records[sid].description.split(" ", 1)[1]
                    meta = json.loads(header_json)
                    org_info.append((sid, meta.get("organism_name")))
                except Exception:
                    org_info.append((sid, None))
        org_df = pd.DataFrame(org_info, columns=["subject", "organism_name"])
        df = df.merge(org_df, on="subject", how="left")

        # Add probable type strain flag
        df["probable_type_strain"] = df["organism_name"].apply(
            lambda x: self.is_probable_type_strain(str(x))
        )
        df["normalized_name"] = df["organism_name"].apply(self.normalize_species)

        # Keep only probable type strains
        df = df[df["probable_type_strain"] == True]

        # Deduplicate organisms (keep highest identity)
        df = df.loc[df.groupby("normalized_name")["identity"].idxmax()].reset_index(drop=True)

        return df

    def find_intersection_with_gold(self, gold_csv, filtered_hits):
        gold_species = set(pd.read_csv(gold_csv)["Organism.Organism Name"].dropna().str.lower())
        df = filtered_hits
        ortho_species = set(df["organism_name"].dropna().str.lower())
        common_species = [o for o in ortho_species if any(g in o for g in gold_species)]
        return df[df["organism_name"].str.lower().isin(common_species)]

    def run(self):
        og_info = self.get_orthodb_ogs()
        if not og_info:
            return
        self.og_id = og_info[0]["id"]
        orthologs = self.fetch_ortholog_sequences()
        self.save_fasta(orthologs, org=True)

        query_fasta = self.fetch_query_sequence()
        query_seq = list(SeqIO.parse(query_fasta, "fasta"))[0]
        self.query_length = len(query_seq.seq)

        out_tsv = os.path.join(TSVS_DIR, f"blast_results_{self.file_name}.tsv")
        self.run_blastp(query_fasta, os.path.join(FASTAS_DIR, self.ortho_file_name), out_tsv)

        filtered_hits = self.filter_blast_results(out_tsv, self.query_length, os.path.join(FASTAS_DIR, self.ortho_file_name))
        gold_df = self.find_intersection_with_gold(GOLD_CSV, filtered_hits)
        not_gold_df = filtered_hits[~filtered_hits["organism_name"].isin(gold_df["organism_name"])]
        not_gold_path = os.path.join(CSVS_DIR, f"notgold_similarity_{self.file_name}.csv")
        not_gold_df.to_csv(not_gold_path, index=False)


# -------------------- RUN --------------------

for name, uid in uniprot_ids.items():
    FUNGALMENACE(uid, name).run()


# -------------------- EXCEL GENERATION --------------------

FEATURES = ["AMR", "Bio", "Hpat", "TH", "RAD", "SF"]
PROTEIN_CATEGORY = {
    "AMR": ["ERG11","CDR1","MDR1","UPC2","TAC1","MRR1"],
    "Bio": ["BCR1","EFG1","TEC1","HWP1","ALS3","NDT80","ERG251","CZF1","FLO8"],
    "Hpat": ["SAP5","PLB1","LAC1","RIM101"],
    "TH": ["HSP90"],
    "RAD": ["RAD51"],
    "SF": ["brlA","abaA","wetA","srr1"],
}

RED_FILL = PatternFill("solid", fgColor="FF6347")
YELLOW_FILL = PatternFill("solid", fgColor="FFD700")
BLUE_FILL = PatternFill("solid", fgColor="87CEFA")


def read_notgold_results(folder):
    csvs = sorted(glob.glob(os.path.join(folder, "notgold_similarity_*.csv")))
    data, organisms, prots = {}, set(), set()
    for path in csvs:
        df = pd.read_csv(path)
        prot = os.path.basename(path).split("notgold_similarity_")[1].replace(".csv", "")
        for _, row in df.iterrows():
            org, val = row.get("organism_name"), row.get("identity")
            if pd.isna(org) or pd.isna(val): 
                continue
            organisms.add(org)
            data[(org, prot)] = val
            prots.add(prot)
    return data, sorted(organisms), sorted(prots)


def build_category_mapping(all_proteins):
    prot_to_cat = {}
    for cat, prots in PROTEIN_CATEGORY.items():
        for p_abbr in prots:
            for prot in all_proteins:
                if p_abbr.lower() in prot.lower():
                    prot_to_cat[prot] = cat
    for prot in all_proteins:
        if prot not in prot_to_cat:
            prot_to_cat[prot] = "Uncategorized"
    grouped = []
    for cat in FEATURES + ["Uncategorized"]:
        grouped.extend([p for p, c in prot_to_cat.items() if c == cat])
    return prot_to_cat, [p for p in grouped if p in all_proteins]


def generate_excel(data, organisms, all_proteins, output_path):
    prot_to_cat, ordered_proteins = build_category_mapping(all_proteins)
    wb = Workbook()
    ws = wb.active
    ws.title = "Probable Type Strains"

    ws.cell(row=1, column=1, value="")
    ws.cell(row=1, column=2, value="")
    col = 3
    for cat in FEATURES + ["Uncategorized"]:
        prots = [p for p, c in prot_to_cat.items() if c == cat]
        for _ in prots:
            cell = ws.cell(row=1, column=col, value=cat)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            col += 1

    ws.cell(row=2, column=1, value="Organism").font = Font(bold=True)
    ws.cell(row=2, column=2, value="A-score").font = Font(bold=True)
    for j, prot in enumerate(ordered_proteins, start=3):
        c = ws.cell(row=2, column=j, value=prot)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")

    summary_col = len(ordered_proteins) + 3
    phylum_col = summary_col + 1
    ws.cell(row=2, column=summary_col, value="r,y,r+y").font = Font(bold=True)
    ws.cell(row=2, column=phylum_col, value="Phylum").font = Font(bold=True)

    red_rows = yellow_rows = blue_rows = 0

    for i, org in enumerate(organisms, start=3):
        ws.cell(row=i, column=1, value=org)
        cat_red = {c:0 for c in FEATURES}
        cat_yellow = {c:0 for c in FEATURES}
        row_red = row_yellow = 0

        for j, prot in enumerate(ordered_proteins, start=3):
            val = data.get((org, prot))
            if val is None: continue
            c = ws.cell(row=i, column=j, value=round(val,1))
            c.alignment = Alignment(horizontal="center")
            cat = prot_to_cat.get(prot, "Uncategorized")
            if val > 75:
                c.fill = RED_FILL; cat_red[cat]+=1; row_red+=1
            elif val > 35:
                c.fill = YELLOW_FILL; cat_yellow[cat]+=1; row_yellow+=1
            else:
                c.fill = BLUE_FILL

        a_score = sum(2 if cat_red[c]>0 else 1 if cat_yellow[c]>0 else 0 for c in FEATURES)
        ws.cell(row=i, column=2, value=a_score).alignment = Alignment(horizontal="center")
        hdr = ws.cell(row=i, column=1)
        if row_red>0: hdr.fill=RED_FILL; red_rows+=1
        elif row_yellow>0: hdr.fill=YELLOW_FILL; yellow_rows+=1
        else: hdr.fill=BLUE_FILL; blue_rows+=1
        ws.cell(row=i, column=summary_col, value=f"{row_red},{row_yellow},{row_red+row_yellow}")
        phylum = get_phylum(org) or "Unknown"
        ws.cell(row=i, column=phylum_col, value=phylum)

    for j in range(1, phylum_col+1):
        ws.column_dimensions[get_column_letter(j)].width = 18 if j>2 else 30
    ws.freeze_panes = "C3"

    total_rows = len(organisms)
    last_row = total_rows + 5
    ws.cell(row=last_row, column=1, value="Summary Statistics").font = Font(bold=True, underline="single")
    ws.cell(row=last_row+1, column=1, value="Total probable type strains"); ws.cell(row=last_row+1, column=2, value=total_rows)
    ws.cell(row=last_row+2, column=1, value="Rows ≥1 red"); ws.cell(row=last_row+2, column=2, value=red_rows)
    ws.cell(row=last_row+3, column=1, value="Rows ≥1 yellow (no red)"); ws.cell(row=last_row+3, column=2, value=yellow_rows)
    ws.cell(row=last_row+4, column=1, value="Completely blue rows"); ws.cell(row=last_row+4, column=2, value=blue_rows)
    ws.cell(row=last_row+6, column=1, value="% with red"); ws.cell(row=last_row+6, column=2, value=round(red_rows/total_rows*100,1))
    ws.cell(row=last_row+7, column=1, value="% with yellow"); ws.cell(row=last_row+7, column=2, value=round(yellow_rows/total_rows*100,1))
    ws.cell(row=last_row+8, column=1, value="% blue"); ws.cell(row=last_row+8, column=2, value=round(blue_rows/total_rows*100,1))

    wb.save(output_path)
    print(f"Excel saved: {output_path}")


data, organisms, prots = read_notgold_results(CSVS_DIR)
generate_excel(data, organisms, prots, EXCEL_PATH)
